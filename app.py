"""
API para procesar archivos XLSB y generar datos JSON.

Esta API permite:
- Subir archivos XLSB
- Listar las hojas disponibles (excluyendo hojas SQL)
- Obtener datos de cada hoja en formato JSON
- Exportar hojas a archivos JSON

Uso:
    uvicorn xlsb_api:app --reload --port 8000
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional, List
from pyxlsb import open_workbook
from pymongo import MongoClient
from dotenv import load_dotenv
import openpyxl
import json
import os
import tempfile
import shutil
import re
import unicodedata
from pathlib import Path
from datetime import datetime

# Cargar variables de entorno
load_dotenv()

# Conexión a MongoDB
MONGODB_URI = os.getenv("MONGODB_URI")
mongo_client = None
db = None

def get_database():
    """Obtiene la conexión a la base de datos MongoDB."""
    global mongo_client, db
    if mongo_client is None:
        mongo_client = MongoClient(MONGODB_URI)
        db = mongo_client.get_database("sena_metas")
    return db

def normalize_collection_name(sheet_name: str) -> str:
    """
    Normaliza el nombre de la hoja para usarlo como nombre de colección.
    Convierte a minúsculas y elimina caracteres especiales.
    """
    # Normalizar caracteres unicode (remover acentos)
    normalized = unicodedata.normalize('NFKD', sheet_name)
    normalized = normalized.encode('ASCII', 'ignore').decode('ASCII')
    # Convertir a minúsculas
    normalized = normalized.lower()
    # Reemplazar espacios y caracteres especiales por guiones bajos
    normalized = re.sub(r'[^a-z0-9]', '_', normalized)
    # Eliminar guiones bajos múltiples
    normalized = re.sub(r'_+', '_', normalized)
    # Eliminar guiones bajos al inicio y final
    normalized = normalized.strip('_')
    return normalized

def is_ejecucion_fpi_file(filename: str) -> bool:
    """Verifica si el archivo es de tipo 'ejecución FPI'."""
    normalized = filename.lower()
    return 'ejecucion fpi' in normalized or 'ejecución fpi' in normalized

def is_metas_file(filename: str) -> bool:
    """Verifica si el archivo es de tipo 'metas SENA'."""
    normalized = filename.lower()
    return 'seguimiento' in normalized and 'metas' in normalized and 'sena' in normalized

app = FastAPI(
    title="XLSB to JSON API",
    description="API para convertir archivos XLSB a JSON",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc",
    openapi_url="/openapi.json"
)

# Configurar CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Directorio temporal para archivos subidos y generados
UPLOAD_DIR = Path(tempfile.gettempdir()) / "xlsb_api_uploads"
OUTPUT_DIR = Path(tempfile.gettempdir()) / "xlsb_api_output"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# Almacén de archivos subidos (en memoria para demo)
uploaded_files = {}


def read_xlsb_sheet(file_path: str, sheet_name: str) -> List[dict]:
    """
    Lee una hoja de un archivo XLSB y retorna los datos como lista de diccionarios.

    Args:
        file_path: Ruta al archivo XLSB
        sheet_name: Nombre de la hoja a leer

    Returns:
        Lista de diccionarios con los datos de la hoja
    """
    data = []
    with open_workbook(file_path) as wb:
        with wb.get_sheet(sheet_name) as sheet:
            rows = list(sheet.rows())
            if not rows:
                return data

            # Primera fila como encabezados
            headers = []
            for cell in rows[0]:
                value = cell.v
                if value is None:
                    value = f"column_{len(headers)}"
                headers.append(str(value).strip())

            # Procesar filas de datos
            for row in rows[1:]:
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        value = cell.v
                        # Convertir valores numéricos
                        if isinstance(value, (int, float)):
                            row_data[headers[i]] = value
                        elif value is not None:
                            row_data[headers[i]] = str(value).strip()
                        else:
                            row_data[headers[i]] = None

                # Solo agregar filas que tengan al menos un valor no nulo
                if any(v is not None for v in row_data.values()):
                    data.append(row_data)

    return data


def get_sheet_names(file_path: str, exclude_sql: bool = True) -> List[str]:
    """
    Obtiene los nombres de las hojas de un archivo XLSB.

    Args:
        file_path: Ruta al archivo XLSB
        exclude_sql: Si es True, excluye hojas con "SQL" en el nombre

    Returns:
        Lista de nombres de hojas
    """
    with open_workbook(file_path) as wb:
        sheets = wb.sheets
        if exclude_sql:
            sheets = [s for s in sheets if 'SQL' not in s.upper()]
        return sheets


def read_xlsx_sheet_metas(file_path: str, sheet_name: str) -> List[dict]:
    """
    Lee una hoja de un archivo XLSX (metas) y extrae solo las columnas de Cupos.

    Args:
        file_path: Ruta al archivo XLSX
        sheet_name: Nombre de la hoja a leer

    Returns:
        Lista de diccionarios con las metas por centro/regional
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Hoja '{sheet_name}' no encontrada en el archivo")

    ws = wb[sheet_name]

    # Buscar fila de encabezados (contiene "Cupos")
    header_row_num = None
    for row_num in range(1, 20):
        row_data = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
        if any('Cupos' in str(cell) if cell else False for cell in row_data):
            header_row_num = row_num
            break

    if not header_row_num:
        wb.close()
        raise ValueError("No se encontró fila de encabezados con 'Cupos'")

    # Leer filas importantes
    fila_categorias = list(ws.iter_rows(min_row=header_row_num-1, max_row=header_row_num-1, values_only=True))[0]
    fila_encabezados = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))[0]

    # Determinar tipo de hoja y mapear columnas de identificación
    # REGIONAL: solo tiene Cód.Reg y Regional (columnas A y B)
    # CTROS: tiene Cód.Reg, Regional, Código Centro y Centro (columnas A, B, C y D)
    is_regional = 'REGIONAL' in sheet_name.upper()

    if is_regional:
        # Hoja "4. FORMACIÓN X REGIONAL": solo 2 columnas de identificación
        col_cod_reg = 0
        col_regional = 1
        col_cod_centro = None
        col_centro = None
    else:
        # Hoja "5. FORMACIÓN X CTROS": 4 columnas de identificación
        col_cod_reg = 0
        col_regional = 1
        col_cod_centro = 2
        col_centro = 3

    # Encontrar columnas de "Cupos" (metas)
    columnas_cupos = []
    mapeo_nombres = {
        'Tecnólogos Regular - Presencial': 'M_TEC_REG_PRE',
        'Tecnólogos Regular - Virtual': 'M_TEC_REG_VIR',
        'Tecnólogos Regular - A Distancia': 'M_TEC_REG_A_D',
        'Tecnólogos CampeSENA': 'M_TEC_CAMPESE',
        'Tecnólogos Full Popular': 'M_TEC_FULL_PO',
        'SubTotal Tecnólogos (E)': 'M_TECNOLOGOS',
        'EDUCACION SUPERIOR (=E)': 'M_EDU_SUPERIO',
        'Operarios Regular': 'M_OPE_REGULAR',
        'Operarios CampeSENA': 'M_OPE_CAMPESE',
        'Operarios Full Popular': 'M_OPE_FULL_PO',
        'SubTotal Operarios (B)': 'M_SUB_TOT_OPE',
        'Auxiliares Regular': 'M_AUX_REGULAR',
        'Auxiliares CampeSENA': 'M_AUX_CAMPESE',
        'Auxiliares Full Popular': 'M_AUX_FULL_PO',
        'SubTotal Auxiliares (A)': 'M_SUB_TOT_AUX',
        'Técnico Laboral Regular - Presencial': 'M_TCO_REG_PRE',
        'Técnico Laboral Regular - Virtual': 'M_TCO_REG_VIR',
        'Técnico Laboral CampeSENA': 'M_TCO_CAMPESE',
        'Técnico Laboral Full Popular': 'M_TCO_FULL_PO',
        'Técnico Laboral Articulación con la Media': 'M_TCO_ART_MED',
        'SubTotal Técnico Laboral (C)': 'M_SUB_TCO_LAB',
        'Profesional Técnico (T)': 'M_PROF_TECNIC',
        'TOTAL FORMACIÓN LABORAL': 'M_TOT_FOR_LAB',
        'TOTAL FORMACION TITULADA': 'M_TOT_FOR_TIT',
        'Complementaria Virtual Sin Bilingüismo': 'M_COM_VIR_SBI',
        'Complementaria Presencial Sin Bilingüismo': 'M_COM_PRE_SBI',
        'Complementaria Bilingüismo Virtual': 'M_COM_BIL_VIR',
        'Complementaria Bilingüismo Presencial': 'M_COM_BIL_PRE',
        'SubTotal Programas Bilingües': 'M_SUB_PRO_BIN',
        'Complementaria CampeSENA': 'M_COM_CAMPESE',
        'Complementaria Full Popular': 'M_COM_FULL_PO',
        'TOTAL COMPLEMENTARIA': 'M_TOT_COMPLEM',
        'TOTAL FORMACIÓN PROFESIONAL': 'M_TOT_PROF_IN'
    }

    # Función auxiliar para normalizar texto (eliminar tildes y comparar)
    def normalizar_texto(texto: str) -> str:
        """Elimina tildes y normaliza texto para comparación."""
        if not texto:
            return ""
        # Normalizar caracteres unicode (eliminar tildes)
        normalized = unicodedata.normalize('NFKD', texto)
        normalized = normalized.encode('ASCII', 'ignore').decode('ASCII')
        return normalized.lower().strip()

    for i, encabezado in enumerate(fila_encabezados):
        if encabezado and 'Cupos' in str(encabezado):
            categoria = fila_categorias[i] if i < len(fila_categorias) and fila_categorias[i] else None

            # Buscar nombre normalizado
            campo_nombre = None
            if categoria:
                categoria_normalizada = normalizar_texto(str(categoria))
                for nombre_original, nombre_normalizado in mapeo_nombres.items():
                    nombre_original_normalizado = normalizar_texto(nombre_original)
                    # Comparar versiones normalizadas (sin tildes)
                    if nombre_original_normalizado in categoria_normalizada:
                        campo_nombre = nombre_normalizado
                        break

            # Si no hay mapeo, crear uno genérico
            if not campo_nombre and categoria:
                campo_nombre = f"M_{normalize_collection_name(str(categoria))}"

            if campo_nombre:
                columnas_cupos.append({
                    'indice': i,
                    'nombre': campo_nombre,
                    'categoria_original': categoria
                })

    # Leer datos
    data_start_row = header_row_num + 1
    resultados = []

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        # Verificar que la fila tenga datos
        if not any(row):
            continue

        registro = {
            'PERIODO': '2025',  # Por defecto, podría extraerse del nombre del archivo
            'COD_REGIONAL': row[col_cod_reg] if col_cod_reg < len(row) else None,
            'REGIONAL': row[col_regional] if col_regional < len(row) else None
        }

        # Solo agregar columnas de centro si existen en esta hoja
        if col_cod_centro is not None and col_centro is not None:
            registro['COD_CENTRO'] = row[col_cod_centro] if col_cod_centro < len(row) else None
            registro['CENTRO'] = row[col_centro] if col_centro < len(row) else None
        else:
            registro['COD_CENTRO'] = None
            registro['CENTRO'] = None

        # Agregar valores de metas
        for col_info in columnas_cupos:
            valor = row[col_info['indice']] if col_info['indice'] < len(row) else None
            # Convertir a número si es posible
            if valor is not None:
                try:
                    valor = float(valor) if isinstance(valor, (int, float)) else 0
                except:
                    valor = 0
            else:
                valor = 0

            registro[col_info['nombre']] = valor

        # Solo agregar si tiene código de regional o centro
        if registro.get('COD_REGIONAL') or registro.get('COD_CENTRO'):
            resultados.append(registro)

    wb.close()
    return resultados


@app.get("/")
async def root():
    """Endpoint raíz con información de la API."""
    return {
        "message": "XLSB to JSON API",
        "version": "1.0.0",
        "endpoints": {
            "POST /upload": "Subir archivo XLSB",
            "GET /files": "Listar archivos subidos",
            "GET /files/{file_id}/sheets": "Listar hojas de un archivo",
            "GET /files/{file_id}/sheets/{sheet_name}": "Obtener datos de una hoja",
            "POST /files/{file_id}/export": "Exportar todas las hojas a JSON",
            "GET /download/{filename}": "Descargar archivo generado"
        }
    }


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """
    Sube un archivo XLSB para procesamiento.
    Si el nombre contiene 'ejecución FPI', guarda los datos en MongoDB.

    Returns:
        ID del archivo y lista de hojas disponibles
    """
    if not file.filename.endswith('.xlsb'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsb")

    # Generar ID único
    file_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    file_path = UPLOAD_DIR / file_id

    # Guardar archivo
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    # Obtener hojas
    try:
        sheets = get_sheet_names(str(file_path))
    except Exception as e:
        os.remove(file_path)
        raise HTTPException(status_code=400, detail=f"Error al leer archivo: {str(e)}")

    # Registrar archivo
    uploaded_files[file_id] = {
        "path": str(file_path),
        "original_name": file.filename,
        "sheets": sheets,
        "uploaded_at": datetime.now().isoformat()
    }

    # Si es archivo de ejecución FPI, guardar en MongoDB
    mongodb_collections = []
    if is_ejecucion_fpi_file(file.filename):
        try:
            database = get_database()
            for sheet_name in sheets:
                # Leer datos de la hoja
                data = read_xlsb_sheet(str(file_path), sheet_name)

                if data:
                    # Crear nombre de colección normalizado
                    collection_name = f"ejecucion_fpi_{normalize_collection_name(sheet_name)}"

                    # Obtener colección y limpiar datos anteriores
                    collection = database[collection_name]
                    collection.delete_many({})

                    # Insertar nuevos datos
                    collection.insert_many(data)

                    mongodb_collections.append({
                        "sheet_name": sheet_name,
                        "collection_name": collection_name,
                        "records_inserted": len(data)
                    })
        except Exception as e:
            # No fallar el upload si MongoDB tiene problemas
            mongodb_collections.append({
                "error": f"Error al guardar en MongoDB: {str(e)}"
            })

    response = {
        "file_id": file_id,
        "original_name": file.filename,
        "sheets": sheets,
        "message": f"Archivo subido exitosamente. {len(sheets)} hojas disponibles (excluyendo SQL)."
    }

    if mongodb_collections:
        response["mongodb_collections"] = mongodb_collections
        response["message"] += f" Datos guardados en {len([c for c in mongodb_collections if 'collection_name' in c])} colecciones de MongoDB."

    return response


@app.post("/upload-metas")
async def upload_metas(file: UploadFile = File(...)):
    """
    Sube un archivo Excel de metas SENA (.xlsx) y guarda las metas en MongoDB.

    El archivo debe contener la hoja "5. FORMACIÓN X CTROS" o "4. FORMACIÓN X REGIONAL"
    con las columnas de "Cupos" que representan las metas.

    Returns:
        Información sobre las metas procesadas y guardadas en MongoDB
    """
    if not file.filename.endswith(('.xlsx', '.xlsb')):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsx o .xlsb")

    if not is_metas_file(file.filename):
        raise HTTPException(
            status_code=400,
            detail="El archivo no parece ser un archivo de metas SENA. Debe contener 'Seguimiento', 'Metas' y 'SENA' en el nombre."
        )

    # Generar ID único
    file_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    file_path = UPLOAD_DIR / file_id

    # Guardar archivo
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        database = get_database()
        mongodb_collections = []

        # Procesar hoja de CENTROS
        try:
            datos_centros = read_xlsx_sheet_metas(str(file_path), "5. FORMACIÓN X CTROS")

            if datos_centros:
                collection_name = "metas_fpi_centros"
                collection = database[collection_name]

                # Limpiar datos anteriores
                collection.delete_many({})

                # Insertar nuevos datos
                collection.insert_many(datos_centros)

                mongodb_collections.append({
                    "sheet_name": "5. FORMACIÓN X CTROS",
                    "collection_name": collection_name,
                    "records_inserted": len(datos_centros)
                })
        except Exception as e:
            mongodb_collections.append({
                "sheet_name": "5. FORMACIÓN X CTROS",
                "error": f"Error al procesar: {str(e)}"
            })

        # Procesar hoja de REGIONAL
        try:
            datos_regional = read_xlsx_sheet_metas(str(file_path), "4. FORMACIÓN X REGIONAL")

            if datos_regional:
                collection_name = "metas_fpi_regional"
                collection = database[collection_name]

                # Limpiar datos anteriores
                collection.delete_many({})

                # Insertar nuevos datos
                collection.insert_many(datos_regional)

                mongodb_collections.append({
                    "sheet_name": "4. FORMACIÓN X REGIONAL",
                    "collection_name": collection_name,
                    "records_inserted": len(datos_regional)
                })
        except Exception as e:
            mongodb_collections.append({
                "sheet_name": "4. FORMACIÓN X REGIONAL",
                "error": f"Error al procesar: {str(e)}"
            })

        # Limpiar archivo temporal
        os.remove(file_path)

        success_count = len([c for c in mongodb_collections if 'collection_name' in c])
        error_count = len([c for c in mongodb_collections if 'error' in c])

        return {
            "file_name": file.filename,
            "processing_date": datetime.now().isoformat(),
            "collections_processed": success_count,
            "errors": error_count,
            "details": mongodb_collections,
            "message": f"Archivo procesado exitosamente. {success_count} colecciones actualizadas."
        }

    except Exception as e:
        # Limpiar archivo en caso de error
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=500, detail=f"Error al procesar archivo: {str(e)}")


@app.get("/files")
async def list_files():
    """Lista todos los archivos subidos."""
    return {
        "files": [
            {
                "file_id": fid,
                "original_name": info["original_name"],
                "sheets": info["sheets"],
                "uploaded_at": info["uploaded_at"]
            }
            for fid, info in uploaded_files.items()
        ]
    }


@app.get("/files/{file_id}/sheets")
async def get_sheets(file_id: str):
    """
    Lista las hojas disponibles en un archivo.

    Args:
        file_id: ID del archivo subido
    """
    if file_id not in uploaded_files:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    return {
        "file_id": file_id,
        "sheets": uploaded_files[file_id]["sheets"]
    }


@app.get("/files/{file_id}/sheets/{sheet_name}")
async def get_sheet_data(
    file_id: str,
    sheet_name: str,
    limit: Optional[int] = Query(None, description="Límite de registros"),
    offset: Optional[int] = Query(0, description="Offset de registros")
):
    """
    Obtiene los datos de una hoja en formato JSON.

    Args:
        file_id: ID del archivo subido
        sheet_name: Nombre de la hoja
        limit: Número máximo de registros a retornar
        offset: Número de registros a saltar
    """
    if file_id not in uploaded_files:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    if sheet_name not in uploaded_files[file_id]["sheets"]:
        raise HTTPException(status_code=404, detail=f"Hoja '{sheet_name}' no encontrada")

    try:
        data = read_xlsb_sheet(uploaded_files[file_id]["path"], sheet_name)

        # Aplicar paginación
        total = len(data)
        if offset:
            data = data[offset:]
        if limit:
            data = data[:limit]

        return {
            "file_id": file_id,
            "sheet_name": sheet_name,
            "total_records": total,
            "returned_records": len(data),
            "offset": offset,
            "limit": limit,
            "data": data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al leer hoja: {str(e)}")


@app.post("/files/{file_id}/export")
async def export_all_sheets(file_id: str):
    """
    Exporta todas las hojas a archivos JSON individuales.

    Args:
        file_id: ID del archivo subido

    Returns:
        Lista de archivos generados con sus URLs de descarga
    """
    if file_id not in uploaded_files:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    exported_files = []
    file_info = uploaded_files[file_id]
    base_name = Path(file_info["original_name"]).stem

    for sheet_name in file_info["sheets"]:
        try:
            data = read_xlsb_sheet(file_info["path"], sheet_name)

            output = {
                "sheet_name": sheet_name,
                "total_records": len(data),
                "generated_at": datetime.now().isoformat(),
                "data": data
            }

            # Guardar archivo
            filename = f"{base_name}_{sheet_name}.json"
            output_path = OUTPUT_DIR / filename

            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(output, f, ensure_ascii=False, indent=2)

            exported_files.append({
                "sheet_name": sheet_name,
                "filename": filename,
                "records": len(data),
                "download_url": f"/download/{filename}"
            })

        except Exception as e:
            exported_files.append({
                "sheet_name": sheet_name,
                "error": str(e)
            })

    return {
        "file_id": file_id,
        "exported_files": exported_files
    }


@app.get("/download/{filename}")
async def download_file(filename: str):
    """
    Descarga un archivo generado.

    Args:
        filename: Nombre del archivo a descargar
    """
    file_path = OUTPUT_DIR / filename

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    return FileResponse(
        path=str(file_path),
        filename=filename,
        media_type="application/json"
    )


@app.delete("/files/{file_id}")
async def delete_file(file_id: str):
    """
    Elimina un archivo subido.

    Args:
        file_id: ID del archivo a eliminar
    """
    if file_id not in uploaded_files:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    # Eliminar archivo físico
    try:
        os.remove(uploaded_files[file_id]["path"])
    except:
        pass

    # Eliminar registro
    del uploaded_files[file_id]

    return {"message": f"Archivo {file_id} eliminado"}


# Endpoint adicional para procesar archivo local (útil para desarrollo)
@app.post("/process-local")
async def process_local_file(
    file_path: str = Query(..., description="Ruta al archivo XLSB local"),
    output_dir: Optional[str] = Query(None, description="Directorio de salida")
):
    """
    Procesa un archivo XLSB local y genera archivos JSON de salida.

    Args:
        file_path: Ruta absoluta al archivo XLSB
        output_dir: Directorio donde guardar los archivos generados
    """
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"Archivo no encontrado: {file_path}")

    if not file_path.endswith('.xlsb'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsb")

    # Directorio de salida
    if output_dir:
        out_path = Path(output_dir)
    else:
        out_path = Path(file_path).parent / "output"

    out_path.mkdir(exist_ok=True)

    # Obtener hojas
    sheets = get_sheet_names(file_path)
    base_name = Path(file_path).stem

    exported_files = []

    for sheet_name in sheets:
        try:
            data = read_xlsb_sheet(file_path, sheet_name)

            output = {
                "sheet_name": sheet_name,
                "total_records": len(data),
                "generated_at": datetime.now().isoformat(),
                "data": data
            }

            # Guardar archivo
            filename = f"{base_name}_{sheet_name}.json"
            output_file = out_path / filename

            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(output, f, ensure_ascii=False, indent=2)

            exported_files.append({
                "sheet_name": sheet_name,
                "filename": filename,
                "output_path": str(output_file),
                "records": len(data)
            })

        except Exception as e:
            exported_files.append({
                "sheet_name": sheet_name,
                "error": str(e)
            })

    return {
        "source_file": file_path,
        "output_directory": str(out_path),
        "sheets_processed": len(sheets),
        "exported_files": exported_files
    }


# =============================================================================
# Endpoints para MongoDB
# =============================================================================

@app.get("/mongodb/collections")
async def list_mongodb_collections():
    """
    Lista todas las colecciones disponibles en MongoDB.
    """
    try:
        database = get_database()
        collections = database.list_collection_names()

        # Filtrar solo colecciones de ejecucion_fpi
        fpi_collections = [c for c in collections if c.startswith("ejecucion_fpi_")]

        collection_info = []
        for collection_name in fpi_collections:
            count = database[collection_name].count_documents({})
            collection_info.append({
                "collection_name": collection_name,
                "document_count": count
            })

        return {
            "total_collections": len(fpi_collections),
            "collections": collection_info
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al conectar con MongoDB: {str(e)}")


@app.get("/mongodb/collections/{collection_name}")
async def get_collection_data(
    collection_name: str,
    limit: Optional[int] = Query(None, description="Límite de registros"),
    offset: Optional[int] = Query(0, description="Offset de registros"),
    sort_by: Optional[str] = Query(None, description="Campo por el cual ordenar"),
    sort_order: Optional[int] = Query(1, description="Orden: 1 ascendente, -1 descendente")
):
    """
    Obtiene los datos de una colección de MongoDB.

    Args:
        collection_name: Nombre de la colección
        limit: Número máximo de registros a retornar
        offset: Número de registros a saltar
        sort_by: Campo para ordenar
        sort_order: 1 para ascendente, -1 para descendente
    """
    try:
        database = get_database()

        # Verificar que la colección existe
        if collection_name not in database.list_collection_names():
            raise HTTPException(status_code=404, detail=f"Colección '{collection_name}' no encontrada")

        collection = database[collection_name]
        total = collection.count_documents({})

        # Construir query
        cursor = collection.find({}, {"_id": 0})

        # Aplicar ordenamiento
        if sort_by:
            cursor = cursor.sort(sort_by, sort_order)

        # Aplicar paginación
        if offset:
            cursor = cursor.skip(offset)
        if limit:
            cursor = cursor.limit(limit)

        data = list(cursor)

        return {
            "collection_name": collection_name,
            "total_records": total,
            "returned_records": len(data),
            "offset": offset,
            "limit": limit,
            "data": data
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener datos: {str(e)}")


@app.get("/mongodb/collections/{collection_name}/schema")
async def get_collection_schema(collection_name: str):
    """
    Obtiene el esquema (campos) de una colección basado en el primer documento.

    Args:
        collection_name: Nombre de la colección
    """
    try:
        database = get_database()

        if collection_name not in database.list_collection_names():
            raise HTTPException(status_code=404, detail=f"Colección '{collection_name}' no encontrada")

        collection = database[collection_name]

        # Obtener primer documento para inferir esquema
        sample = collection.find_one({}, {"_id": 0})

        if not sample:
            return {
                "collection_name": collection_name,
                "fields": [],
                "message": "Colección vacía"
            }

        fields = []
        for key, value in sample.items():
            field_type = type(value).__name__
            fields.append({
                "field_name": key,
                "field_type": field_type
            })

        return {
            "collection_name": collection_name,
            "total_fields": len(fields),
            "fields": fields
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener esquema: {str(e)}")


@app.get("/mongodb/collections/{collection_name}/search")
async def search_collection(
    collection_name: str,
    field: str = Query(..., description="Campo a buscar"),
    value: str = Query(..., description="Valor a buscar"),
    exact: bool = Query(False, description="Búsqueda exacta o parcial"),
    limit: Optional[int] = Query(100, description="Límite de registros")
):
    """
    Busca documentos en una colección por un campo específico.

    Args:
        collection_name: Nombre de la colección
        field: Campo donde buscar
        value: Valor a buscar
        exact: Si es True, búsqueda exacta. Si es False, búsqueda parcial (regex)
        limit: Número máximo de resultados
    """
    try:
        database = get_database()

        if collection_name not in database.list_collection_names():
            raise HTTPException(status_code=404, detail=f"Colección '{collection_name}' no encontrada")

        collection = database[collection_name]

        # Construir query
        if exact:
            # Intentar convertir a número si es posible
            try:
                numeric_value = float(value)
                query = {field: numeric_value}
            except ValueError:
                query = {field: value}
        else:
            # Búsqueda parcial con regex (case insensitive)
            query = {field: {"$regex": value, "$options": "i"}}

        cursor = collection.find(query, {"_id": 0}).limit(limit)
        data = list(cursor)

        return {
            "collection_name": collection_name,
            "search_field": field,
            "search_value": value,
            "exact_match": exact,
            "results_count": len(data),
            "data": data
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en búsqueda: {str(e)}")


@app.delete("/mongodb/collections/{collection_name}")
async def delete_collection(collection_name: str):
    """
    Elimina una colección de MongoDB.

    Args:
        collection_name: Nombre de la colección a eliminar
    """
    try:
        database = get_database()

        if collection_name not in database.list_collection_names():
            raise HTTPException(status_code=404, detail=f"Colección '{collection_name}' no encontrada")

        database[collection_name].drop()

        return {
            "message": f"Colección '{collection_name}' eliminada exitosamente"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al eliminar colección: {str(e)}")


@app.get("/mongodb/collections/{collection_name}/aggregate")
async def aggregate_collection(
    collection_name: str,
    group_by: str = Query(..., description="Campo por el cual agrupar"),
    aggregate_field: Optional[str] = Query(None, description="Campo numérico para agregar"),
    operation: str = Query("count", description="Operación: count, sum, avg, min, max")
):
    """
    Realiza agregaciones básicas en una colección.

    Args:
        collection_name: Nombre de la colección
        group_by: Campo por el cual agrupar
        aggregate_field: Campo numérico para operaciones sum/avg/min/max
        operation: Tipo de operación (count, sum, avg, min, max)
    """
    try:
        database = get_database()

        if collection_name not in database.list_collection_names():
            raise HTTPException(status_code=404, detail=f"Colección '{collection_name}' no encontrada")

        collection = database[collection_name]

        # Construir pipeline de agregación
        if operation == "count":
            pipeline = [
                {"$group": {"_id": f"${group_by}", "value": {"$sum": 1}}},
                {"$sort": {"value": -1}}
            ]
        elif operation in ["sum", "avg", "min", "max"]:
            if not aggregate_field:
                raise HTTPException(
                    status_code=400,
                    detail=f"Se requiere 'aggregate_field' para la operación '{operation}'"
                )
            op_map = {
                "sum": "$sum",
                "avg": "$avg",
                "min": "$min",
                "max": "$max"
            }
            pipeline = [
                {"$group": {"_id": f"${group_by}", "value": {op_map[operation]: f"${aggregate_field}"}}},
                {"$sort": {"value": -1}}
            ]
        else:
            raise HTTPException(status_code=400, detail=f"Operación '{operation}' no soportada")

        results = list(collection.aggregate(pipeline))

        # Formatear resultados
        formatted_results = [
            {group_by: r["_id"], operation: r["value"]}
            for r in results
        ]

        return {
            "collection_name": collection_name,
            "group_by": group_by,
            "operation": operation,
            "aggregate_field": aggregate_field,
            "results_count": len(formatted_results),
            "data": formatted_results
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en agregación: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
