"""
Script para analizar la estructura del archivo Excel de metas
"""
import openpyxl
import sys

# Cargar el archivo
archivo_path = r'C:\ws\sena\data\seguimiento_metas\sena-metas-app\data\aprendices\Seguimiento a Metas SENA 2025 V5 111125.xlsx'
wb = openpyxl.load_workbook(archivo_path, data_only=True, read_only=True)

# Listar hojas
print('=' * 80)
print('HOJAS DISPONIBLES:')
print('=' * 80)
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f'{i}. {sheet_name}')

# Buscar la hoja de FORMACIÓN X CTROS
target_sheet = None
for sheet_name in wb.sheetnames:
    if 'FORMACIÓN' in sheet_name.upper() and 'CTROS' in sheet_name.upper():
        target_sheet = sheet_name
        break

if target_sheet:
    print(f'\n{"=" * 80}')
    print(f'ANALIZANDO HOJA: {target_sheet}')
    print('=' * 80)
    ws = wb[target_sheet]

    # Obtener dimensiones
    print(f'\nDimensiones: {ws.max_row} filas x {ws.max_column} columnas')

    # Buscar la fila de encabezados (buscar "Cupos" en las primeras 15 filas)
    print('\n' + '=' * 80)
    print('BUSCANDO FILA DE ENCABEZADOS...')
    print('=' * 80)

    header_row_num = None
    header_row = None

    for row_num in range(1, 16):
        row_data = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
        # Buscar si tiene "Cupos" en alguna celda
        has_cupos = any('Cupos' in str(cell) if cell else False for cell in row_data)
        if has_cupos:
            header_row_num = row_num
            header_row = row_data
            print(f'\n¡Encabezados encontrados en fila {row_num}!')
            break

        # Mostrar contenido de primeras filas
        first_non_null = next((str(cell) for cell in row_data if cell), 'vacía')
        print(f'Fila {row_num}: {first_non_null[:80]}...')

    if not header_row:
        print('\nERROR: No se encontró fila con encabezados "Cupos"')
        wb.close()
        sys.exit(1)

    # Mostrar encabezados
    print('\n' + '=' * 80)
    print(f'ENCABEZADOS (Fila {header_row_num}):')
    print('=' * 80)
    for i, cell_value in enumerate(header_row, 1):
        if cell_value:
            print(f'Col {openpyxl.utils.get_column_letter(i)} ({i:2d}): {cell_value}')

    # Buscar columnas con 'Cupos'
    print('\n' + '=' * 80)
    print('COLUMNAS CON "Cupos" (METAS):')
    print('=' * 80)
    cupos_columns = []
    for i, cell_value in enumerate(header_row, 1):
        if cell_value and 'Cupos' in str(cell_value):
            cupos_columns.append((i, openpyxl.utils.get_column_letter(i), cell_value))
            print(f'Col {openpyxl.utils.get_column_letter(i)} ({i:2d}): {cell_value}')

    print(f'\nTotal de columnas de Cupos encontradas: {len(cupos_columns)}')

    # Mostrar primeras filas de datos
    print('\n' + '=' * 80)
    print('PRIMERAS 3 FILAS DE DATOS:')
    print('=' * 80)
    data_start_row = header_row_num + 1
    for i, row in enumerate(ws.iter_rows(min_row=data_start_row, max_row=data_start_row + 2, values_only=True), data_start_row):
        print(f'\nFila {i}:')
        # Mostrar solo primeras 10 columnas
        for j, val in enumerate(row[:10], 1):
            if j-1 < len(header_row) and header_row[j-1]:
                print(f'  {header_row[j-1]}: {val}')

else:
    print('\nERROR: No se encontró hoja con FORMACIÓN X CTROS')

wb.close()
print('\n' + '=' * 80)
print('ANÁLISIS COMPLETADO')
print('=' * 80)
