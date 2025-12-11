"""
Script para mapear columnas de Cupos (metas) con sus categorías.
Analiza tanto la hoja de CENTROS como la de REGIONAL para mostrar diferencias.
"""
import openpyxl
import json

archivo_path = r'C:\ws\sena\data\seguimiento_metas\sena-metas-app\data\aprendices\Seguimiento a Metas SENA 2025 V5 -25112025.xlsx'
wb = openpyxl.load_workbook(archivo_path, data_only=True, read_only=True)

print('=' * 100)
print('ANÁLISIS DE ESTRUCTURA DE ARCHIVOS DE METAS')
print('=' * 100)
print()

# Analizar ambas hojas
hojas_a_analizar = [
    '5. FORMACIÓN X CTROS',
    '4. FORMACIÓN X REGIONAL'
]

for sheet_name in hojas_a_analizar:
    print('\n' + '=' * 100)
    print(f'ANALIZANDO HOJA: {sheet_name}')
    print('=' * 100)

    ws = wb[sheet_name]

    # Buscar fila de encabezados
    header_row_num = None
    for row_num in range(1, 20):
        row_data = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
        if any('Cupos' in str(cell) if cell else False for cell in row_data):
            header_row_num = row_num
            break

    if not header_row_num:
        print(f'ERROR: No se encontró fila de encabezados en {sheet_name}')
        continue

    # Leer las filas importantes
    fila_categorias = list(ws.iter_rows(min_row=header_row_num-1, max_row=header_row_num-1, values_only=True))[0]
    fila_encabezados = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))[0]

    # Determinar número de columnas de identificación
    is_regional = 'REGIONAL' in sheet_name.upper()
    num_id_cols = 2 if is_regional else 4

    # Primeras columnas de identificación
    print(f'\nCOLUMNAS DE IDENTIFICACIÓN (primeras {num_id_cols} columnas):')
    print('-' * 100)
    for i in range(num_id_cols):
        print(f'Col {openpyxl.utils.get_column_letter(i+1)}: ')
        print(f'  Fila {header_row_num-1} (Categoría): {fila_categorias[i] if i < len(fila_categorias) else None}')
        print(f'  Fila {header_row_num} (Encabezado): {fila_encabezados[i] if i < len(fila_encabezados) else None}')
        print()

    # Encontrar todas las columnas de "Cupos" y su mapeo
    print('MAPEO DE COLUMNAS "Cupos" (METAS):')
    print('-' * 100)

    mapeo_cupos = []
    for i, encabezado in enumerate(fila_encabezados):
        if encabezado and 'Cupos' in str(encabezado):
            categoria = fila_categorias[i] if i < len(fila_categorias) else None

            # Información de ejecución y porcentaje
            ejecucion_header = fila_encabezados[i+1] if i+1 < len(fila_encabezados) else None
            porcentaje_header = fila_encabezados[i+2] if i+2 < len(fila_encabezados) else None

            mapeo_cupos.append({
                'columna': openpyxl.utils.get_column_letter(i+1),
                'indice': i,
                'categoria': categoria,
                'header_cupos': encabezado,
                'header_ejecucion': ejecucion_header,
                'header_porcentaje': porcentaje_header
            })

            print(f'\n{len(mapeo_cupos)}. Columna {openpyxl.utils.get_column_letter(i+1)} (índice {i}):')
            print(f'   Categoría: {categoria}')
            print(f'   Cupos: {encabezado}')
            print(f'   Ejecución: {ejecucion_header}')
            print(f'   Porcentaje: {porcentaje_header}')

    print(f'\n✓ Total de columnas de metas en {sheet_name}: {len(mapeo_cupos)}')

    # Mostrar diferencias clave
    if is_regional:
        print(f'\n⚠️  NOTA: Esta hoja tiene solo {num_id_cols} columnas de identificación.')
        print(f'   Las columnas de Cupos empiezan en la columna C (índice 2)')
    else:
        print(f'\n⚠️  NOTA: Esta hoja tiene {num_id_cols} columnas de identificación.')
        print(f'   Las columnas de Cupos empiezan en la columna E (índice 4)')

wb.close()

print('\n' + '=' * 100)
print('RESUMEN DE DIFERENCIAS ENTRE HOJAS')
print('=' * 100)
print("""
Diferencias clave entre las hojas:

1. HOJA "4. FORMACIÓN X REGIONAL":
   - 2 columnas de identificación: COD_REGIONAL (A), REGIONAL (B)
   - Primera columna de Cupos: Columna C (índice 2)
   - En MongoDB: COD_CENTRO y CENTRO se establecen como null

2. HOJA "5. FORMACIÓN X CTROS":
   - 4 columnas de identificación: COD_REGIONAL (A), REGIONAL (B), COD_CENTRO (C), CENTRO (D)
   - Primera columna de Cupos: Columna E (índice 4)
   - En MongoDB: Todos los campos de identificación tienen valores

3. MAPEO DE CATEGORÍAS:
   - El sistema normaliza el texto eliminando tildes
   - "Tecnologos" (Excel sin tilde) → "Tecnólogos" (mapeo con tilde) ✓
   - Permite matching robusto independiente de acentos
""")
